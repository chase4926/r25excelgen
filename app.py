#!/usr/bin/python

"""
Delta College R25 Excel Tool

TODO:
Don't combine reservations if there is a class in between them

----
Requirements:
- Python 2.7
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from datetime import time
from datetime import *
from collections import defaultdict
import re
import math
import glob
import cocos
from cocos.audio.pygame.mixer import Sound
from cocos.audio.pygame import mixer
import pyglet


def format_time(t):
  time_string = t.strftime("%I:%M %p")
  if time_string[0] == "0":
    time_string = time_string[1:]
  return time_string

def time_to_datetime(t):
  current = date.today()
  return datetime.combine(current, t)

def time_between(t1, t2):
  return abs(time_to_datetime(t1) - time_to_datetime(t2))

def unicode_to_time(u):
  new_time = datetime.strptime(u.encode('ASCII', 'ignore'), '%I:%M %p')
  return time(new_time.hour, new_time.minute, new_time.second)

def resource_common_name(resource):
  return {
    'Laptop wifi': 'Laptop',
    'Computer / Dell Laptop': 'Laptop',
    'Laptop Wireless Cart #1 (20)': 'Cart',
    'Laptop Wireless Cart #2 (20)': 'Cart',
    'Laptop Wireless Cart #3 (20)': 'Cart',
    'Clickers 25': 'Clickers',
    'Clickers 52': 'Clickers-50',
    'Wireless Presenter': 'Presenter',
  }.get(resource, resource)

class Audio(Sound):
  def __init__(self, audio_file):
    # As stated above, we initialize the super class with the audio file we passed in
    super(Audio, self).__init__(audio_file)

class Event:
  def __init__(self):
    self.space = None
    self.resource = None
    self.start = None
    self.end = None
    self.delivery_window = None
    self.pickup_window = None
    self.delivery_done = False
    self.pickup_done = False
  
  def add_resource(self, resource):
    if self.resource == None:
      self.resource = [resource]
    else:
      if not resource in self.resource:
        self.resource.append(resource)
  
  def set_start(self, t):
    #Set start time
    if type(t) is time:
      self.start = t
    else:
      self.start = unicode_to_time(t)
  
  def time_difference(self, event):
    # returns the smallest difference in time between self and event
    time_list = list()
    for t1 in (self.start, self.end):
      for t2 in (event.start, event.end):
        time_list.append(abs(time_to_datetime(t1) - time_to_datetime(t2)))
    return sorted(time_list)[0]
  
  def set_end(self, t):
    # Set end time
    if type(t) is time:
      self.end = t
    else:
      self.end = unicode_to_time(t)
  
  def __lt__(a, b):
    if a.start < b.start:
      return True
    elif a.start == b.start:
      if a.end <= b.end:
        return True
      else:
        return False
    else:
      return False
  
  def __gt__(a, b):
    if a.start > b.start:
      return True
    elif a.start == b.start:
      if a.end > b.end:
        return True
      else:
        return False
    else:
      return False
  
  def get_pickup_time(self):
    if len(self.pickup_window) == 1:
      if self.pickup_window[0] == 'OPEN':
        return 'OPEN'
      else:
        return "@%s" % format_time(self.pickup_window[0])
    else:
      return "%s-%s" % (format_time(self.pickup_window[0]), format_time(self.pickup_window[1]))
  
  def get_delivery_time(self):
    if len(self.delivery_window) == 1:
      if self.delivery_window[0] == 'OPEN':
        return 'OPEN'
      else:
        return "@%s" % format_time(self.delivery_window[0])
    else:
      return "%s-%s" % (format_time(self.delivery_window[0]), format_time(self.delivery_window[1]))
  
  def __repr__(self):
    return "%s - %s, %s, %s" % (format_time(self.start), format_time(self.end), self.space, self.resource)


class EventBook:
  def load_workbook(self, filename):
    self.workbook = load_workbook(filename)
    self.sheet = self.workbook.active
    self.rooms = self.get_room_events()
    print("There are %i total reservations." % len(self.get_reservations(self.rooms)))
    self.reservations = self.get_reservations(self.combine_reservations(self.rooms))
    print("After being combined, there are %i reservations." % len(self.reservations))
    self.reservations = sorted(self.reservations)
    for event in self.reservations:
      self.process_delivery_time(event)
      self.process_pickup_time(event)

  def get_room_events(self):
    rooms = defaultdict(list)
    last_event = None
    for row in tuple(self.sheet.rows):
      if row[0].value != None:
        if type(row[0].value) is time:
          if row[5].value != None:
            event = Event()
            event.set_start(row[0].value) # Start datetime.time
            event.set_end(row[1].value) # End datetime.time
            event.space = self.format_space(row[5].value.encode('ASCII', 'ignore')) # Space
            if row[6].value != None:
              event.add_resource(row[6].value.encode('ASCII', 'ignore')) # Resource
            rooms[event.space].append(event)
            # Set last_event in case there are additional resources
            last_event = event
      else:
        # Check if there was an event prior to this
        if last_event != None:
          if row[6].value != None:
            #Add resource to last event
            last_event.add_resource(row[6].value.encode('ASCII', 'ignore'))
          else:
            # Trailed off the end of the list
            last_event = None
    return rooms
  
  def get_reservations(self, rooms):
    reservations = list()
    for room in rooms:
      for event in rooms[room]:
        if event.resource != None:
          reservations.append(event)
    return reservations
  
  def copy_rooms(self, rooms):
    new_rooms = defaultdict(list)
    for room in rooms:
      new_rooms[room] = rooms[room][:]
    return new_rooms
  
  def combine_reservations(self, r):
    # Input a rooms dict, outputs a new dict with combined events
    rooms = self.copy_rooms(r)
    for room in rooms:
      combined = True
      while(combined):
        combined = False
        for e1 in rooms[room]:
          for e2 in rooms[room]:
            if e1 != e2 and not combined:
              #Don't compare the event to itself
              if (e1.resource != None) and (e1.resource == e2.resource) and (e1.time_difference(e2) < timedelta(hours=2)):
                combined = True
                if e1 < e2: # Checks if e1 takes place before e2
                  e1.end = e2.end
                else:
                  e1.start = e2.start
                rooms[room].remove(e2)
    return rooms

  def process_pickup_time(self, event):
    events = self.rooms[event.space]
    after_events = list()
    for e in events:
      if event.end < e.start:
        after_events.append(e)
    if len(after_events) > 0:
      start_event = after_events[0]
      start_time = start_event.start
      if event.time_difference(start_event) < timedelta(minutes=15):
        event.pickup_window = [start_time]
      else:
        event.pickup_window = [event.end, start_time]
    else:
      event.pickup_window = ['OPEN']
  
  def process_delivery_time(self, event):
    events = self.rooms[event.space]
    events = events[0:events.index(event)]
    if len(events) > 0:
      end_event = events[-1]
      end_time = end_event.end
      if event.time_difference(end_event) < timedelta(minutes=15):
        event.delivery_window = [end_time]
      else:
        event.delivery_window = [end_time, event.start]
    else:
      event.delivery_window = ['OPEN']
  
  def format_space(self, space):
    result = space.split('_', 1)[1]
    return result.split(' ', 1)[0]
  
  def get_resource_column(self, resource):
    return {
      'Laptop wifi': 'D',
      'Computer / Dell Laptop': 'D',
      'Laptop Wireless Cart #1 (20)': 'E',
      'Laptop Wireless Cart #2 (20)': 'E',
      'Laptop Wireless Cart #3 (20)': 'E',
      'Clickers 25': 'F',
      'Clickers 52': 'F',
      'Wireless Presenter': 'G',
    }.get(resource, 'H') # Misc
  
  def save_workbook(self, template_filename='template.xlsx'):
    template_book = load_workbook(template_filename)
    template = template_book.active
    
    i = 2 # Row to start on
    for event in self.reservations:
      # Write event details to row
      template["B%i" % i] = event.space
      template["K%i" % i] = format_time(event.start)
      template["M%i" % i] = format_time(event.end)
      for resource in event.resource:
        resource_col = self.get_resource_column(resource)
        if resource_col == 'H':
          template["H%i" % i] = resource
        else:
          template["%s%i" % (resource_col, i)] = 'X'
      template["J%i" % i] = event.get_delivery_time()
      template["N%i" % i] = event.get_pickup_time()
      # Move to the next row
      i += 1
    
    # Save Excel Document for processed events
    red_color = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')
    for i in range(len(self.reservations)):
      if i % 2 == 0:
        for n in range(15):
          template[("%s%i") % (chr(n+65), i + 2)].fill = red_color
    
    template_book.save("%s.xlsx" % (datetime.now() + timedelta(days=1)).strftime("%b-%d"))

  def get_current_deliveries(self, current_time, td=2):
    # Returns a list of tuples. (td = time difference in hours)
    # tuples contain (event, priority) where priority is 1-100 with 100 being urgent
    result = list()
    for event in self.reservations:
      if len(event.delivery_window) == 1:
        # Either OPEN or @
        if event.delivery_window[0] == 'OPEN':
          # Deliver before event.start
          if current_time < event.start and time_between(current_time, event.start) < timedelta(hours=td):
            result.append( (event, 1) ) # FIXME: Priority
        else:
          # Deliver right at [0]
          if current_time < event.delivery_window[0] and time_between(current_time, event.delivery_window[0]) < timedelta(minutes=60):
            result.append( (event, 1) ) # FIXME: Priority
      else:
        # Deliver after [0] and before [1]
        if current_time > event.delivery_window[0] and current_time < event.delivery_window[1] and time_between(current_time, event.delivery_window[1]) < timedelta(hours=td):
          result.append( (event, 1) ) # FIXME: Priority
    return result
  
  def get_current_pickups(self, current_time, td=2):
    # Returns a list of tuples. (td = time difference in hours)
    # tuples contain (event, priority) where priority is 1-100 with 100 being urgent
    result = list()
    for event in self.reservations:
      if len(event.pickup_window) == 1:
        #Either OPEN or @
        if event.pickup_window[0] == 'OPEN':
          # Pickup after event.end
          if current_time > event.end and time_between(event.end, current_time) < timedelta(hours=td):
            result.append( (event, 1) ) # FIXME: Priority
        else:
          # Pickup right at [0]
          if current_time < event.pickup_window[0] and time_between(current_time, event.pickup_window[0]) < timedelta(minutes=60):
            result.append( (event, 1) ) # FIXME: Priority
      else:
        # Pickup after [0] and before [1]
        if (current_time > event.pickup_window[0] and current_time < event.pickup_window[1]) or time_between(current_time, event.pickup_window[0]) < timedelta(minutes=60):
          result.append( (event, 1) ) # FIXME: Priority
    return result


# Gui Code


class UpdateEvent(cocos.actions.InstantAction):
  def start(self):
    self.target.update_events()


class EventWindow(cocos.layer.Layer):
  is_event_handler = True
  
  def __init__(self):
    super(EventWindow, self).__init__()
    self.wb = EventBook()
    self.wb.load_workbook('reservations.xlsx')
    self.wb.save_workbook() # Save the next days schedule automatically
    self.font = 'Arial'
    # Deliveries
    delivery_label = cocos.text.Label(
      'Deliveries',
      font_name = self.font,
      font_size = 24,
      anchor_x = 'center', anchor_y = 'center'
    )
    delivery_label.position = 100, 690
    self.add(delivery_label)
    # Pickups
    pickup_label = cocos.text.Label(
      'Pickups',
      font_name = self.font,
      font_size = 24,
      anchor_x = 'center', anchor_y = 'center'
    )
    pickup_label.position = 1190, 690
    self.add(pickup_label)
    # Updated at ...
    self.time_label = cocos.text.Label(
      'Not updated!',
      font_name = self.font,
      font_size = 24,
      anchor_x = 'center', anchor_y = 'center'
    )
    self.time_label.position = 640, 690
    self.add(self.time_label)
    # Delivery labels
    self.delivery_slots = list()
    for i in range(25):
      label = cocos.text.Label(
        '',
        font_name = self.font,
        font_size = 14,
        anchor_x = 'left', anchor_y = 'center'
      )
      label.position = 20, 640 - i*24
      self.delivery_slots.append([label, None])
      self.add(label)
    # Pickup labels
    self.pickup_slots = list()
    for i in range(25):
      label = cocos.text.Label(
        '',
        font_name = self.font,
        font_size = 14,
        anchor_x = 'right', anchor_y = 'center'
      )
      label.position = 1260, 640 - i*24
      self.pickup_slots.append([label, None])
      self.add(label)
    self.do(cocos.actions.Repeat(UpdateEvent() + cocos.actions.Delay(60)))
    self.audio_assets = self.load_audio_assets()
    self.promode = False
    self.promode_last_time = datetime.now().time()
    self.promode_count = 0
    self.promode_order = ('first_blood.wav', 'double_kill.wav', 'megakill.wav', 'ultrakill.wav', 'Monsterkill_F.wav', 'godlike.wav')
  
  def load_audio_assets(self, folder='assets'):
    result = dict()
    for filename in glob.glob("%s/*.wav" % folder):
      filename = filename.replace('\\', '/')
      result[filename.split("%s/"%folder)[1]] = Audio(filename)
    return result
  
  def done_event(self):
    if self.promode:
      if self.promode_count == 0 or time_between(self.promode_last_time, datetime.now().time()) > timedelta(seconds=30):
        self.promode_count = 1
        self.audio_assets[self.promode_order[self.promode_count-1]].play()
      else:
        if self.promode_count < len(self.promode_order):
          self.promode_count += 1
        self.audio_assets[self.promode_order[self.promode_count-1]].play()
      self.promode_last_time = datetime.now().time()
  
  def update_events(self):
    current_time = datetime.now().time()#.replace(hour=10).replace(minute=31)
    self.time_label.element.text = "Updated at %s" % format_time(current_time)
    # Clear old text
    for label_event in self.delivery_slots:
      label_event[0].element.text = ''
      label_event[0].element.color = (255, 255, 255, 255)
      label_event[1] = None
    for label_event in self.pickup_slots:
      label_event[0].element.text = ''
      label_event[0].element.color = (255, 255, 255, 255)
      label_event[1] = None
    # Add new text
    for i, event in enumerate(self.wb.get_current_deliveries(current_time)):
      delivery_time = event[0].get_delivery_time()
      if delivery_time == 'OPEN':
        delivery_time = "OPEN-%s" % format_time(event[0].start)
      self.delivery_slots[i][1] = event[0]
      if event[0].delivery_done:
        self.delivery_slots[i][0].element.color = (255, 0, 0, 255)
      self.delivery_slots[i][0].element.text = "%s | %s | %s" % (
        delivery_time, event[0].space,
        ",".join([resource_common_name(a) for a in event[0].resource]))
    for i, event in enumerate(self.wb.get_current_pickups(current_time)):
      self.pickup_slots[i][1] = event[0]
      if event[0].pickup_done:
        self.pickup_slots[i][0].element.color = (255, 0, 0, 255)
      self.pickup_slots[i][0].element.text = "%s | %s | %s" % (
        event[0].get_pickup_time(), event[0].space,
        ",".join([resource_common_name(a) for a in event[0].resource]))
  
  def mouse_y_to_label_i(self, y, label_y=640, label_spacing=24):
    # Converts mouse y coordinate to label index
    return int(math.floor((y-label_y-(label_spacing/2))/(label_spacing*1.0))*-1)-1
  
  def on_key_press(self, key, mod):
    if key == 65479: # F10
      self.promode = not self.promode
  
  def on_mouse_press(self, x, y, button, mod):
    i = self.mouse_y_to_label_i(y)
    if button == 1 and 0 <= i < len(self.delivery_slots):
      if x < 640:
        #Deliveries
        if self.delivery_slots[i][1] != None:
          if self.delivery_slots[i][1].delivery_done:
            self.delivery_slots[i][1].delivery_done = False
            self.delivery_slots[i][0].element.color = (255, 255, 255, 255)
          else:
            self.done_event()
            self.delivery_slots[i][1].delivery_done = True
            self.delivery_slots[i][0].element.color = (255, 0, 0, 255)
      else:
        #Pickups
        if self.pickup_slots[i][1] != None:
          if self.pickup_slots[i][1].pickup_done:
            self.pickup_slots[i][1].pickup_done = False
            self.pickup_slots[i][0].element.color = (255, 255, 255, 255)
          else:
            self.done_event()
            self.pickup_slots[i][1].pickup_done = True
            self.pickup_slots[i][0].element.color = (255, 0, 0, 255)

# ---

cocos.director.director.init(width=1280, height=720, caption="R25 Excel Gen")
mixer.init()
main_scene = cocos.scene.Scene(EventWindow())
cocos.director.director.run(main_scene)


